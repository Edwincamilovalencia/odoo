# Importaciones necesarias para el manejo de archivos, logs y modelos de Odoo
from odoo import api, fields, models
import base64
import csv
import zipfile
import html as html_escape
import logging
from io import BytesIO, StringIO
import xml.etree.ElementTree as ET

# Intentar importar xlrd para soporte de archivos .xls antiguos
try:
    import xlrd  # para soporte de .xls
except Exception:
    xlrd = None

# Logger para registrar errores y advertencias
_logger = logging.getLogger(__name__)

# Extiende el modelo politico.carga para agregar la previsualización de archivos cargados
class PartidoCargaPreview(models.Model):
    _inherit = 'politico.carga'

    # Campo HTML que almacena la previsualización generada del archivo
    preview_html = fields.Html(string='Previsualización', compute='_compute_preview_html', store=False)

    @api.onchange('upload_file', 'file_type', 'filename')
    def _onchange_preview_html(self):
        # Se ejecuta cuando cambia el archivo, tipo o nombre para actualizar la previsualización
        for rec in self:
            rec.preview_html = rec._generate_preview_html()

    @api.depends('upload_file', 'file_type', 'filename')
    def _compute_preview_html(self):
        # Calcula la previsualización cuando cambian los campos relevantes
        for rec in self:
            rec.preview_html = rec._generate_preview_html()

    # -------- Métodos auxiliares ---------
    def _get_file_content(self):
        # Obtiene el contenido binario del archivo cargado, decodificando base64 si es necesario
        try:
            b64data = self.with_context(bin_size=False).upload_file or self.upload_file
            # Si el archivo es string base64
            if isinstance(b64data, str) and b64data:
                try:
                    return base64.b64decode(b64data)
                except Exception:
                    # Si hay caracteres extraños, limpiar y decodificar
                    cleaned = ''.join(ch for ch in b64data if ch.isalnum() or ch in '+/=\r\n')
                    return base64.b64decode(cleaned)
            # Si el archivo ya es bytes
            elif isinstance(b64data, (bytes, bytearray)):
                try:
                    sample = bytes(b64data)[:16]
                    if all(chr(c) in 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=\r\n' for c in sample):
                        return base64.b64decode(bytes(b64data))
                except Exception:
                    pass
                return bytes(b64data)
            # Si el archivo está guardado como adjunto
            if self.id:
                att = self.env['ir.attachment'].search([
                    ('res_model', '=', 'politico.carga'),
                    ('res_id', '=', self.id),
                    ('res_field', '=', 'upload_file'),
                ], limit=1)
                if att and att.datas:
                    return base64.b64decode(att.with_context(bin_size=False).datas)
        except Exception as e:
            _logger.error('Error obteniendo contenido: %s', e)
        return None

    def _detect_kind(self, content):
        # Detecta el tipo de archivo (csv, xls, xlsx) usando extensión y contenido
        ext = (self.file_type or '').lower()
        # Si no hay tipo explícito, intenta por extensión del nombre
        if not ext and self.filename and '.' in (self.filename or ''):
            ext = self.filename.rsplit('.', 1)[-1].lower()
        if ext not in {'csv', 'xls', 'xlsx'}:
            return None

        # Para xlsx, verifica si es un archivo zip válido
        if ext == 'xlsx':
            try:
                if zipfile.is_zipfile(BytesIO(content)):
                    return 'xlsx'
            except Exception:
                pass
        # Para xls, verifica el magic number
        if ext == 'xls' and content[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
            return 'xls'
        # Si el magic no coincide, igual retorna ext; el llamador puede hacer fallback a csv
        return ext

    def _render_table(self, rows, header=None):
        # Genera una tabla HTML estilizada para mostrar la previsualización de los datos
        def esc(x):
            return '' if x is None else html_escape.escape(str(x))
        if not rows and not header:
            return '<div class="alert alert-warning">No hay datos para mostrar.</div>'
        # Estilos CSS para la tabla y celdas
        table_css = (
            "width: 100%; border-collapse: collapse; font-size: 14px; table-layout: auto;"
            "font-family: 'Segoe UI', Arial, sans-serif; background: #fff;"
        )
        th_css = (
            "background: #f3f3f3; color: #222; font-weight: 600; padding: 10px 8px; border: 1px solid #e0e0e0; text-align: left;"
        )
        td_css = (
            "padding: 8px 8px; border: 1px solid #e0e0e0; overflow-wrap: anywhere; white-space: normal;"
        )
        container_css = (
            "width: 100%; max-width: 100%; box-sizing: border-box; display: block;"
            "max-height: 420px; border-radius: 6px; border: 1px solid #e0e0e0; background: #fafbfc; margin-top: 8px; padding: 0; overflow: auto;"
        )
        parts = [f'<div style="{container_css}">', f'<table style="{table_css}">']
        if header:
            parts.append('<thead><tr>')
            for c in header:
                parts.append(f'<th style="{th_css}">{esc(c)}</th>')
            parts.append('</tr></thead>')
        parts.append('<tbody>')
        for r in rows:
            parts.append('<tr>')
            for c in r:
                parts.append(f'<td style="{td_css}">{esc(c)}</td>')
            parts.append('</tr>')
        parts.append('</tbody></table></div>')
        return ''.join(parts)

    def _parse_csv(self, content, max_rows=20, max_cols=15):
        # Parsea un archivo CSV y devuelve filas y cabecera para previsualización
        text = None
        # Intenta decodificar el archivo en varios formatos comunes
        for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
            try:
                text = content.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            text = content.decode('utf-8', errors='replace')
        # Detecta el delimitador automáticamente
        delimiter = ','
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|:^~')
            delimiter = dialect.delimiter
        except Exception:
            counts = {d: sample.count(d) for d in [',', ';', '\t', '|', ':', '^', '~']}
            delimiter = max(counts, key=counts.get)
        rows = []
        reader = csv.reader(StringIO(text), delimiter=delimiter)
        for i, row in enumerate(reader):
            if i >= max_rows + 1:
                break
            if row:
                rows.append([str(c) if c is not None else '' for c in row[:max_cols]])
        header = None
        # Si la primera fila parece cabecera, la separa
        if rows and any(str(c).strip() for c in rows[0]) and (len(rows) > 1 and len(rows[0]) > 1):
            header = rows.pop(0)
        return rows, header

    def _generate_preview_html(self):
        try:
            if not self.upload_file:
                return

            content = self._get_file_content()
            if not content:
                return '<div class="alert alert-danger">No se pudo leer el archivo.</div>'

            kind = self._detect_kind(content)
            if not kind:
                return '<div class="alert alert-warning">Solo se previsualizan archivos CSV o Excel (.csv, .xls, .xlsx).</div>'

            max_rows, max_cols = 20, 15

            # Priorizar motores nativos para Excel antes que pandas (más fiable)
            if kind == 'xlsx':
                # 1) Intento con openpyxl si está disponible
                try:
                    import openpyxl
                    try:
                        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
                        try:
                            ws = wb.active
                            rows_iter = ws.iter_rows(min_row=1, max_row=max_rows + 1, max_col=max_cols, values_only=True)
                            all_rows = [['' if v is None else v for v in row] for row in rows_iter]
                        finally:
                            wb.close()
                        if not any(any(str(c).strip() for c in r) for r in all_rows):
                            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=False)
                            try:
                                ws = wb.active
                                rows_iter = ws.iter_rows(min_row=1, max_row=max_rows + 1, max_col=max_cols, values_only=True)
                                all_rows = [['' if v is None else v for v in row] for row in rows_iter]
                            finally:
                                wb.close()
                        header = None
                        if all_rows and any(str(c).strip() for c in all_rows[0]):
                            header = all_rows.pop(0)
                        return self._render_table(all_rows, header)
                    except zipfile.BadZipFile:
                        rows, header = self._parse_csv(content, max_rows=max_rows, max_cols=max_cols)
                        return self._render_table(rows, header)
                    except Exception:
                        # Si openpyxl falla, continuamos con fallback nativo XML
                        pass
                except Exception:
                    # openpyxl no disponible; seguimos con fallback XML
                    pass

                # 2) Fallback nativo: parsear XLSX (ZIP + XML) para una vista previa básica
                try:
                    zf = zipfile.ZipFile(BytesIO(content))
                except Exception as e:
                    return f'<div class="alert alert-danger">Archivo XLSX inválido: {html_escape.escape(str(e))}</div>'

                # shared strings (puede no existir)
                shared_strings = []
                try:
                    with zf.open('xl/sharedStrings.xml') as f:
                        sst = ET.parse(f).getroot()
                        # namespace simplificado
                        for si in sst.iter():
                            if si.tag.endswith('}si'):
                                # puede tener t directo o rich text (r/t)
                                txt = ''
                                for t in si.iter():
                                    if t.tag.endswith('}t') and t.text:
                                        txt += t.text
                                shared_strings.append(txt)
                except KeyError:
                    shared_strings = []
                except Exception:
                    shared_strings = []

                # localizar la primera hoja
                target_sheet = 'xl/worksheets/sheet1.xml'
                try:
                    with zf.open('xl/workbook.xml') as f:
                        wbxml = ET.parse(f).getroot()
                    relmap = {}
                    try:
                        with zf.open('xl/_rels/workbook.xml.rels') as f:
                            rels = ET.parse(f).getroot()
                            for rel in rels.iter():
                                if rel.tag.endswith('}Relationship'):
                                    relmap[rel.get('Id')] = rel.get('Target')
                    except Exception:
                        relmap = {}
                    first_rid = None
                    for sh in wbxml.iter():
                        if sh.tag.endswith('}sheet') and (first_rid is None):
                            first_rid = sh.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    if first_rid and first_rid in relmap:
                        tgt = relmap[first_rid]
                        if not tgt.startswith('xl/'):
                            target_sheet = 'xl/' + tgt
                        else:
                            target_sheet = tgt
                except Exception:
                    pass

                # helpers de columna (A->1)
                def col_to_idx(col_letters: str) -> int:
                    n = 0
                    for ch in col_letters:
                        if 'A' <= ch <= 'Z':
                            n = n * 26 + (ord(ch) - 64)
                        elif 'a' <= ch <= 'z':
                            n = n * 26 + (ord(ch.upper()) - 64)
                    return n

                # parsear filas/celdas
                rows_map = {}
                try:
                    with zf.open(target_sheet) as f:
                        sh = ET.parse(f).getroot()
                        for row_el in sh.iter():
                            if not row_el.tag.endswith('}row'):
                                continue
                            r_index = int(row_el.get('r') or 0)
                            if r_index == 0 or r_index > (max_rows + 1):
                                continue
                            row_vals = {}
                            for c in row_el:
                                if not c.tag.endswith('}c'):
                                    continue
                                ref = c.get('r') or ''  # p.ej. A1
                                # columna
                                col_letters = ''.join(ch for ch in ref if ch.isalpha())
                                col = col_to_idx(col_letters)
                                t = c.get('t')  # s, b, str, inlineStr, etc
                                v_text = ''
                                if t == 'inlineStr':
                                    # <is><t>...</t></is>
                                    for tnode in c.iter():
                                        if tnode.tag.endswith('}t') and tnode.text:
                                            v_text += tnode.text
                                else:
                                    v_el = next((child for child in c if child.tag.endswith('}v')), None)
                                    if v_el is not None and v_el.text is not None:
                                        v_text = v_el.text
                                        if t == 's':
                                            try:
                                                idx = int(v_text)
                                                v_text = shared_strings[idx] if 0 <= idx < len(shared_strings) else v_text
                                            except Exception:
                                                pass
                                row_vals[col] = v_text
                            rows_map[r_index] = row_vals
                except KeyError:
                    return '<div class="alert alert-danger">No se encontró la primera hoja en el XLSX.</div>'
                except Exception as e:
                    return f'<div class="alert alert-danger">Error leyendo XLSX (XML): {html_escape.escape(str(e))}</div>'

                # construir matriz con huecos vacíos
                max_row = min(max(rows_map.keys() or [0]), max_rows + 1)
                max_col = 0
                for r in range(1, max_row + 1):
                    cols = rows_map.get(r, {})
                    if cols:
                        max_col = max(max_col, min(max(cols.keys()), max_cols))
                all_rows = []
                for r in range(1, max_row + 1):
                    cols = rows_map.get(r, {})
                    row_list = []
                    for c in range(1, max_col + 1):
                        row_list.append(cols.get(c, ''))
                    all_rows.append(row_list)

                header = None
                if all_rows and any(str(c).strip() for c in all_rows[0]):
                    header = all_rows.pop(0)
                return self._render_table(all_rows, header)

            if kind == 'xls':
                if not xlrd:
                    return '<div class="alert alert-danger">Para previsualizar .xls se requiere xlrd==1.2.0.</div>'
                try:
                    book = xlrd.open_workbook(file_contents=content)
                    sheet = book.sheet_by_index(0)
                    all_rows = []
                    limit = min(max_rows + 1, sheet.nrows)
                    for r in range(limit):
                        row = []
                        for c in range(min(max_cols, sheet.ncols)):
                            v = sheet.cell_value(r, c)
                            row.append('' if v is None else v)
                        all_rows.append(row)
                    header = None
                    if all_rows and any(str(c).strip() for c in all_rows[0]):
                        header = all_rows.pop(0)
                    return self._render_table(all_rows, header)
                except Exception as e:
                    return f'<div class="alert alert-danger">Error leyendo XLS: {html_escape.escape(str(e))}</div>'

            # Para CSV: pandas opcional y luego fallback propio
            pd = None
            try:
                import pandas as _pd  # type: ignore
                pd = _pd
            except Exception:
                pd = None

            if kind == 'csv' and pd:
                try:
                    df = pd.read_csv(BytesIO(content), sep=None, engine='python', encoding='utf-8', nrows=max_rows)
                    if df is not None and not df.empty:
                        header = [str(c) for c in list(df.columns)[:max_cols]]
                        rows = df.fillna('').astype(str).values.tolist()
                        rows = [r[:max_cols] for r in rows]
                        return self._render_table(rows, header)
                except Exception as e:
                    _logger.debug('Pandas falló con CSV, usando fallback: %s', e)

            if kind == 'csv':
                rows, header = self._parse_csv(content, max_rows=max_rows, max_cols=max_cols)
                return self._render_table(rows, header)

            return '<div class="alert alert-warning">No se pudo procesar el archivo.</div>'

        except Exception as e:
            _logger.exception('Error en previsualización')
            return f'<div class="alert alert-danger">Error al previsualizar: {html_escape.escape(str(e))}</div>'
